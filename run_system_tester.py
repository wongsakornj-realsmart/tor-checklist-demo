import os
import sys

from backend.services.ocr_service import extract_text_from_file
from backend.services.ai_engine import generate_tor_checklist
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

download_dir = r"C:\Users\womgsakorn_j\Downloads"
os.makedirs(download_dir, exist_ok=True)

test_files = [
    (r"D:\ทหารผ่านศึก\TOR (1).pdf", os.path.join(download_dir, "TOR (1)_Checklist.xlsx"))
]

def create_excel_report(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOR Checklist"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "TH Sarabun PSK"
    title_font = Font(name=font_family, size=18, bold=True, color="000000")
    header_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=16, bold=False, color="000000")
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    meta = data.get("metadata", {})
    project_name = meta.get("project_name", "") or "โครงการจัดซื้อจัดจ้าง"
    client_name = meta.get("client_name", "") or "หน่วยงานเจ้าของโครงการ"
    dateline = meta.get("dateline", "") or "-"
    
    ws.cell(row=1, column=1, value=f"ชื่อโครงการ (Project Name): {project_name}").font = title_font
    ws.cell(row=2, column=1, value=f"ชื่อลูกค้า (Client Name): {client_name}").font = title_font
    ws.cell(row=3, column=1, value=f"Dateline: {dateline}").font = title_font
    
    headers = [
        "Status", "ลำดับ", "หมวดหมู่หลัก", "หัวข้อย่อย",
        "ข้อกำหนด / รายละเอียด (Requirement / Details)",
        "ชื่อเอกสารที่ใช้ยื่น", "รายละเอียดที่ต้องระบุ",
        "Comply?", "หมายเหตุ (Remarks)"
    ]
    
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[5].height = 35
    
    checklist = data.get("checklist", [])
    current_row = 6
    for item in checklist:
        row_values = [
            item.get("Status", ""),
            item.get("ลำดับ", ""),
            item.get("หมวดหมู่หลัก", ""),
            item.get("หัวข้อย่อย", ""),
            item.get("ข้อกำหนด / รายละเอียด (Requirement / Details)", ""),
            item.get("ชื่อเอกสารที่ใช้ยื่น", ""),
            item.get("รายละเอียดที่ต้องระบุ", ""),
            item.get("Comply?", "False"),
            item.get("หมายเหตุ (Remarks)", "")
        ]
        
        for c_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=str(val))
            cell.font = data_font
            cell.border = thin_border
            
            if current_row % 2 == 1:
                cell.fill = alt_fill
                
            if c_idx in [1, 2, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif c_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
        ws.row_dimensions[current_row].height = 28
        current_row += 1
        
    col_widths = {1: 12, 2: 10, 3: 25, 4: 25, 5: 60, 6: 35, 7: 50, 8: 12, 9: 20}
    for c_idx, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = width
        
    wb.save(output_path)
    print(f"[Success] Generated Excel file: {output_path} with {len(checklist)} items.")

for pdf_path, excel_path in test_files:
    print(f"\n==================================================")
    print(f"Testing PDF: {pdf_path}")
    print(f"==================================================")
    text = extract_text_from_file(pdf_path)
    print(f"Extracted {len(text)} chars of clean text.")
    data = generate_tor_checklist(text)
    create_excel_report(data, excel_path)
