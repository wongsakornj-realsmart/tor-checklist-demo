import pdfplumber
import pandas as pd
import re
import os
import openpyxl
from pythainlp.util import normalize

def fix_spelling(text):
    if not text:
        return text
    text = str(text)
    
    # CRITICAL FIX: Strip illegal XML control characters to prevent openpyxl IllegalCharacterError crashes
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    
    text = normalize(text)
    
    # 1. Regex universal OCR fixes
    # Fix double vowels/tones (e.g. ประสิทธิิภาพ -> ประสิทธิภาพ)
    text = re.sub(r'([ิีึืุู่้๊๋์])\1+', r'\1', text)
    # Fix misplaced Sara I/U after two consonants with Sara E (e.g. ดำเนนิ -> ดำเนิน)
    text = re.sub(r'เ([ก-ฮ])([ก-ฮ])([ิีึืุูั])([่้๊๋์]?)', r'เ\1\3\4\2', text)
    
    # 2. Dictionary of common OCR spelling errors in technical documents
    common_fixes = {
        'ผยู้ น': 'ผู้ยื่น',
        'ผยู้น': 'ผู้ยื่น',
        'กำหน ด': 'กำหนด',
        'ขอบเข ต': 'ขอบเขต',
        'เคื่รอง': 'เครื่อง',
        'ทเี่ ': 'ที่เ',
        'เจ้าหน้าทเี่ ': 'เจ้าหน้าที่',
        'เวบไซต์': 'เว็บไซต์',
        'เว็ป': 'เว็บ',
        'อินเทอรเน็ต': 'อินเทอร์เน็ต',
        'ซอฟตแวร': 'ซอฟต์แวร์',
        'ฮารดแวร': 'ฮาร์ดแวร์',
        'เซิรฟเวอร': 'เซิร์ฟเวอร์',
        'คอมพิวเตอร': 'คอมพิวเตอร์',
        'เวอรชัน': 'เวอร์ชัน',
        'แพลตฟอรม': 'แพลตฟอร์ม',
        'อัปเดต': 'อัปเดต',
        'พาสเวิรด': 'พาสเวิร์ด',
        'ไฟล': 'ไฟล์',
        'ออนไลน': 'ออนไลน์',
        'ฟังกชัน': 'ฟังก์ชัน',
        'แอปพลิเคชน': 'แอปพลิเคชัน',
        'ระบบปฎิบัติการ': 'ระบบปฏิบัติการ',
        'ปฎิบัติ': 'ปฏิบัติ',
        'สิทธ': 'สิทธิ',
        'เปี่ลยน': 'เปลี่ยน',
        'เปี่ลย': 'เปลี่ยน',
        'เปีรยบ': 'เปรียบ',
        'เสิรม': 'เสริม',
        'เจิรญ': 'เจริญ',
        'ืรอ': 'รือ',
        'ีรยม': 'รียม',
        'คณุ': 'คุณ',
        'บตั': 'บัต',
        'จดั': 'จัด',
        'ไมพ่': 'ไม่พ',
        'ดงั': 'ดัง',
        'เหุตผล': 'เหตุผล',
        'วัตถุประสงค ์': 'วัตถุประสงค์',
        '\uf0a0': '', # Remove private use bullet points
    }
    for wrong, right in common_fixes.items():
        text = text.replace(wrong, right)
    return text

def generate_excel_from_data(rows: list, template_path: str, output_path: str, metadata: dict = None):
    """
    Populates the 9-column TOR Checklist template with extracted/AI-generated rows.
    
    New template structure (updated June 2026):
    - Row 1: Label "ชื่อโครงการ" (col A) + project name (col B)
    - Row 2: Label "ชื่อลูกค้า" (col A) + client name (col B)
    - Row 3: Label "Dateline" (col A) + dateline (col B)
    - Row 4: (blank separator)
    - Row 5: 9-column headers (Status, ลำดับ, หมวดหมู่หลัก, ...)
    - Row 6+: Checklist data rows
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Write Metadata to Rows 1-3 (Column B = column 2)
    if metadata and isinstance(metadata, dict):
        project_name = metadata.get('project_name', '')
        client_name = metadata.get('client_name', '')
        dateline = metadata.get('dateline', '')
        
        if project_name:
            ws.cell(row=1, column=2, value=fix_spelling(project_name))
        if client_name:
            ws.cell(row=2, column=2, value=fix_spelling(client_name))
        if dateline:
            ws.cell(row=3, column=2, value=fix_spelling(dateline))

    # Write Checklist data starting at Row 6 (after Row 5 headers)
    start_row = 6
    for idx, row in enumerate(rows):
        ws.cell(row=start_row + idx, column=1, value=row.get('Status', ''))
        ws.cell(row=start_row + idx, column=2, value=row.get('ลำดับ', ''))
        ws.cell(row=start_row + idx, column=3, value=fix_spelling(row.get('หมวดหมู่หลัก', '')))
        ws.cell(row=start_row + idx, column=4, value=fix_spelling(row.get('หัวข้อย่อย', '')))
        ws.cell(row=start_row + idx, column=5, value=fix_spelling(row.get('ข้อกำหนด / รายละเอียด (Requirement / Details)', '')))
        ws.cell(row=start_row + idx, column=6, value=fix_spelling(row.get('ชื่อเอกสารที่ใช้ยื่น', '')))
        ws.cell(row=start_row + idx, column=7, value=fix_spelling(row.get('รายละเอียดที่ต้องระบุ', '')))
        ws.cell(row=start_row + idx, column=8, value=row.get('Comply?', 'False'))
        ws.cell(row=start_row + idx, column=9, value=row.get('หมายเหตุ (Remarks)', ''))

    wb.save(output_path)
    print(f'Successfully generated {output_path} with {len(rows)} rows (9 columns). Metadata: project="{metadata.get("project_name", "") if metadata else ""}"')
    return output_path

if __name__ == '__main__':
    # Legacy execution support for testing
    print("Extract TOR module loaded successfully. Use generate_excel_from_data for 9-column generation.")
