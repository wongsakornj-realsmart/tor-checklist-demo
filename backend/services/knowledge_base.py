"""
Knowledge Base Builder for TOR Checklist System.

Reads sample TOR documents and sample TOR Checklists to build a structured
knowledge base that teaches the AI how to extract and format checklist items.
"""
import os
import json
import pdfplumber
import openpyxl

# Flexible Root Path Detection
BASE_DIR = r'D:\CBD\TORChecklist' if os.path.exists(r'D:\CBD\TORChecklist') else os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

SAMPLE_TOR_DIR = os.path.join(BASE_DIR, 'ตัวอย่าง TOR')
SAMPLE_CHECKLIST_DIR = os.path.join(BASE_DIR, 'ตัวอย่าง TOR Checklist')
CACHE_PATH = os.path.join(BASE_DIR, 'tor_knowledge_cache.json')

# In-memory cache to avoid rebuilding every request
_knowledge_cache = None


def _extract_pdf_text(pdf_path: str, max_pages: int = 5) -> str:
    """Extracts text from a PDF file (first N pages)."""
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:max_pages]:
                t = page.extract_text()
                if t and t.strip():
                    text += t + "\n"
    except Exception as e:
        print(f"[KnowledgeBase] Error reading PDF {pdf_path}: {e}")
    return text.strip()


def _extract_checklist_patterns(xlsx_path: str, max_rows: int = 500) -> list:
    """
    Reads a sample TOR Checklist Excel file and extracts row patterns
    showing how TOR requirements map to checklist columns.
    Returns a list of dicts representing sample rows.
    """
    patterns = []
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        # Detect header row (find the row containing 'Status')
        header_row = None
        for r in range(1, min(ws.max_row + 1, 10)):
            for c in range(1, min(ws.max_column + 1, 12)):
                v = ws.cell(r, c).value
                if v and 'Status' in str(v):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            return patterns

        # Read data rows after header
        data_start = header_row + 1
        for r in range(data_start, min(data_start + max_rows, ws.max_row + 1)):
            row_data = {}
            col_names = ['Status', 'ลำดับ', 'หมวดหมู่หลัก', 'หัวข้อย่อย',
                         'ข้อกำหนด / รายละเอียด (Requirement / Details)',
                         'ชื่อเอกสารที่ใช้ยื่น', 'รายละเอียดที่ต้องระบุ',
                         'Comply?', 'หมายเหตุ (Remarks)']
            has_data = False
            for c_idx, col_name in enumerate(col_names, 1):
                v = ws.cell(r, c_idx).value
                if v is not None:
                    row_data[col_name] = str(v).strip()
                    if col_name not in ('Status', 'Comply?', 'หมายเหตุ (Remarks)') and str(v).strip():
                        has_data = True
                else:
                    row_data[col_name] = ''
            if has_data:
                patterns.append(row_data)

    except Exception as e:
        print(f"[KnowledgeBase] Error reading Checklist {xlsx_path}: {e}")
    return patterns


def _extract_section_headers(text: str) -> list:
    """
    Extracts common section headers from TOR documents.
    Returns a list of detected section names.
    """
    import re
    headers = []
    # Common TOR section patterns in Thai
    patterns = [
        r'(?:^|\n)\s*[๐-๙\d]+[\.\)]\s*(.+)',           # "1. ความเป็นมา" or "๑. ความเป็นมา"
        r'(?:^|\n)\s*ข้อ\s*[๐-๙\d]+\s*(.+)',              # "ข้อ 1 ..."
        r'(?:^|\n)\s*หมวด(?:ที่)?\s*[๐-๙\d]+\s*(.+)',     # "หมวดที่ 1 ..."
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text)
        for m in matches:
            clean = m.strip()[:80]
            if clean and len(clean) > 3:
                headers.append(clean)
    return list(dict.fromkeys(headers))[:30]  # deduplicate, keep order, max 30


def build_tor_knowledge_base(force_rebuild: bool = False) -> dict:
    """
    Builds or loads the TOR Knowledge Base from sample files.
    Returns a dict with:
      - 'tor_structures': list of section headers from sample TOR docs
      - 'checklist_examples': list of sample checklist row patterns
      - 'category_names': unique category names found in samples
      - 'document_types': unique document types found in samples
    """
    global _knowledge_cache

    # Return in-memory cache if available
    if _knowledge_cache and not force_rebuild:
        return _knowledge_cache

    # Try loading from disk cache
    if not force_rebuild and os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH, 'r', encoding='utf-8') as f:
                _knowledge_cache = json.load(f)
            print(f"[KnowledgeBase] Loaded cache from {CACHE_PATH}")
            return _knowledge_cache
        except Exception as e:
            print(f"[KnowledgeBase] Cache load failed: {e}, rebuilding...")

    print("[KnowledgeBase] Building Knowledge Base from sample files...")
    kb = {
        'tor_structures': [],
        'checklist_examples': [],
        'category_names': set(),
        'document_types': set()
    }

    # --- Step 1: Read sample TOR PDFs ---
    if os.path.isdir(SAMPLE_TOR_DIR):
        for fname in sorted(os.listdir(SAMPLE_TOR_DIR)):
            if fname.lower().endswith('.pdf'):
                fpath = os.path.join(SAMPLE_TOR_DIR, fname)
                text = _extract_pdf_text(fpath, max_pages=5)
                if text:
                    headers = _extract_section_headers(text)
                    kb['tor_structures'].append({
                        'source': fname,
                        'sections': headers,
                        'sample_text': text[:1500]  # Keep only first 1500 chars as reference
                    })
                    print(f"  [TOR] {fname}: {len(headers)} sections extracted")
                else:
                    print(f"  [TOR] {fname}: scanned PDF (no extractable text)")
    else:
        print(f"[KnowledgeBase] Sample TOR directory not found: {SAMPLE_TOR_DIR}")

    # --- Step 2: Read sample TOR Checklist Excel files ---
    if os.path.isdir(SAMPLE_CHECKLIST_DIR):
        for fname in sorted(os.listdir(SAMPLE_CHECKLIST_DIR)):
            if fname.lower().endswith('.xlsx'):
                fpath = os.path.join(SAMPLE_CHECKLIST_DIR, fname)
                rows = _extract_checklist_patterns(fpath, max_rows=500)
                if rows:
                    kb['checklist_examples'].append({
                        'source': fname,
                        'rows': rows
                    })
                    # Collect unique categories and document types
                    for row in rows:
                        cat = row.get('หมวดหมู่หลัก', '')
                        if cat:
                            kb['category_names'].add(cat)
                        doc = row.get('ชื่อเอกสารที่ใช้ยื่น', '')
                        if doc:
                            kb['document_types'].add(doc)
                    print(f"  [Checklist] {fname}: {len(rows)} pattern rows extracted")
    else:
        print(f"[KnowledgeBase] Sample Checklist directory not found: {SAMPLE_CHECKLIST_DIR}")

    # Convert sets to lists for JSON serialization
    kb['category_names'] = sorted(list(kb['category_names']))
    kb['document_types'] = sorted(list(kb['document_types']))

    # Save to disk cache
    try:
        with open(CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(kb, f, ensure_ascii=False, indent=2)
        print(f"[KnowledgeBase] Saved cache to {CACHE_PATH}")
    except Exception as e:
        print(f"[KnowledgeBase] Failed to save cache: {e}")

    _knowledge_cache = kb
    return kb


def get_knowledge_prompt_section(kb: dict) -> str:
    """
    Converts the Knowledge Base into a concise prompt section
    to inject into the AI's system prompt (Dynamic RAG).
    """
    sections = []

    # 1. Common TOR section structure
    all_section_names = set()
    for tor in kb.get('tor_structures', []):
        for s in tor.get('sections', []):
            all_section_names.add(s)
    if all_section_names:
        sections.append(
            "หัวข้อที่พบบ่อยในเอกสาร TOR ของหน่วยงานราชการไทย:\n" +
            ", ".join(sorted(all_section_names)[:25])
        )

    # 2. Common category names
    cats = kb.get('category_names', [])
    if cats:
        sections.append(
            "หมวดหมู่หลักที่ใช้จัดกลุ่มข้อกำหนด:\n" +
            ", ".join(cats[:20])
        )

    # 3. Common document types
    docs = kb.get('document_types', [])
    if docs:
        sections.append(
            "ชื่อเอกสารที่ใช้ยื่นที่พบบ่อย:\n" +
            ", ".join(docs[:20])
        )

    # 4. Sample checklist rows (pick 3 diverse examples from different files)
    examples = kb.get('checklist_examples', [])
    sample_rows = []
    for ex in examples[:5]:
        rows = ex.get('rows', [])
        if rows:
            # Pick first row and a mid-section row to show diversity
            sample_rows.append(rows[0])
            if len(rows) > 10:
                sample_rows.append(rows[len(rows)//2])
    if sample_rows:
        # Limit to 5 diverse examples to keep prompt manageable
        sections.append(
            "ตัวอย่างรูปแบบการเขียนรายการ Checklist (ใช้เป็นแนวทางรูปแบบเท่านั้น ห้ามลอกเนื้อหา):\n" +
            json.dumps(sample_rows[:5], ensure_ascii=False, indent=2)
        )

    return "\n\n".join(sections)


if __name__ == '__main__':
    kb = build_tor_knowledge_base(force_rebuild=True)
    print(f"\n=== Knowledge Base Summary ===")
    print(f"TOR structures: {len(kb['tor_structures'])} files")
    print(f"Checklist examples: {len(kb['checklist_examples'])} files")
    print(f"Category names: {len(kb['category_names'])}")
    print(f"Document types: {len(kb['document_types'])}")
    print(f"\n=== RAG Prompt Preview ===")
    print(get_knowledge_prompt_section(kb))
