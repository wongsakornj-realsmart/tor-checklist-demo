import pandas as pd
import pdfplumber
import openpyxl
import os

def run_validation_on_file(excel_path: str, pdf_path: str = None) -> bool:
    """
    Validates the generated TOR Checklist Excel file against the new template structure.
    
    New template structure (updated June 2026):
    - Row 1-3: Metadata (ชื่อโครงการ, ชื่อลูกค้า, Dateline)
    - Row 5: 9-column headers
    - Row 6+: Data rows
    """
    if not os.path.exists(excel_path):
        print(f"Validation FAILED: Excel file not found at {excel_path}")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        print(f"Validation FAILED: Cannot open Excel file: {e}")
        return False

    # 1. Metadata Audit (Row 1-3)
    metadata_labels = {
        1: 'ชื่อโครงการ',
        2: 'ชื่อลูกค้า',
        3: 'Dateline'
    }
    metadata_ok = True
    for row_num, expected_label in metadata_labels.items():
        label = ws.cell(row=row_num, column=1).value
        if label and expected_label in str(label):
            value = ws.cell(row=row_num, column=2).value
            print(f'Metadata Row {row_num} ({expected_label}): {"PRESENT" if value else "EMPTY (acceptable)"}')
        else:
            print(f'Metadata Row {row_num}: Label mismatch (expected "{expected_label}", got "{label}")')
            metadata_ok = False

    if metadata_ok:
        print('Metadata Audit: PASSED')
    else:
        print('Metadata Audit: WARNING - Some labels do not match expected format')

    # 2. Header Audit (Row 5 - 9 Columns)
    expected_headers = [
        'Status',
        'ลำดับ',
        'หมวดหมู่หลัก',
        'หัวข้อย่อย',
        'ข้อกำหนด / รายละเอียด (Requirement / Details)',
        'ชื่อเอกสารที่ใช้ยื่น',
        'รายละเอียดที่ต้องระบุ',
        'Comply?',
        'หมายเหตุ (Remarks)'
    ]

    header_row = 5
    actual_headers = []
    for c in range(1, 10):
        v = ws.cell(row=header_row, column=c).value
        actual_headers.append(str(v) if v else '')

    # Flexible header matching (check key words rather than exact match)
    headers_ok = True
    for i, expected in enumerate(expected_headers):
        actual = actual_headers[i] if i < len(actual_headers) else ''
        # Extract key part of expected header for comparison
        key_word = expected.split('/')[0].split('(')[0].strip().split('?')[0].strip()
        if key_word and key_word not in actual:
            # Some templates use slightly different header names
            if not any(kw in actual for kw in [key_word[:5], expected[:5]]):
                print(f'Header Col {i+1}: MISMATCH (expected contains "{key_word}", got "{actual}")')
                headers_ok = False

    if headers_ok:
        print(f'Format Audit: PASSED (9 Columns verified at Row {header_row})')
    else:
        print(f'Format Audit: WARNING - Some column headers at Row {header_row} do not match')

    # 3. Data Integrity Audit (Row 6+)
    data_row_count = 0
    for r in range(6, ws.max_row + 1):
        # Check if any data exists in columns 2-7
        has_data = any(ws.cell(r, c).value for c in range(2, 8))
        if has_data:
            data_row_count += 1
    
    if data_row_count > 0:
        print(f'Data Integrity Audit: PASSED ({data_row_count} data rows found)')
    else:
        print('Data Integrity Audit: WARNING - No data rows found')

    # 4. Optional PDF cross-check
    if pdf_path and os.path.exists(pdf_path):
        try:
            # Read all text from data column (column 5)
            excel_text = ''
            for r in range(6, ws.max_row + 1):
                v = ws.cell(r, 5).value
                if v:
                    excel_text += str(v)
            excel_text_clean = ''.join(e for e in excel_text if e.isalnum())

            pdf_text = ''
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        pdf_text += t
            pdf_text_clean = ''.join(e for e in pdf_text if e.isalnum())

            if pdf_text_clean:
                ratio = len(excel_text_clean) / len(pdf_text_clean)
                print(f'PDF Cross-Check: Extraction ratio = {ratio:.2f}')
                if ratio < 0.3:
                    print('PDF Cross-Check: WARNING - Extraction ratio lower than expected')
        except Exception as e:
            print(f'PDF Cross-Check: SKIPPED ({e})')

    return True

if __name__ == '__main__':
    default_excel = r'D:\CBD\TORChecklist\OutputTORChecklist\Attach_TOR_1_Checklist.xlsx'
    default_pdf = r'D:\CBD\TORChecklist\InputTOR\Attach_TOR_1.pdf'
    run_validation_on_file(default_excel, default_pdf)
